import os
import sys
from openpyxl.utils import get_column_letter


def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath('.')
    return os.path.join(base_path, relative_path)

def encontrar_coluna(cabecalho, pagina_clientes):
    for col_idx, cell in enumerate(pagina_clientes[1]):  # Primeira linha é o cabeçalho
        if cell.value == cabecalho:
            return col_idx + 1
    return None

def adicionar_coluna_envio(pagina_clientes):
    nova_coluna_index = pagina_clientes.max_column + 1
    coluna_letra = get_column_letter(nova_coluna_index)
    pagina_clientes[f'{coluna_letra}1'].value = 'Enviado'  # Coloca o cabeçalho na nova coluna
    return nova_coluna_index

def resetar_status_envio(pagina_clientes, coluna_envio):
    coluna_envio = encontrar_coluna("Enviado", pagina_clientes)
    if coluna_envio is not None:
        for linha in pagina_clientes.iter_rows(min_row=2):
            linha[coluna_envio - 1].value = None


def verificar_numeros_enviados(pagina_clientes, coluna_envio):
    if coluna_envio is None:
       coluna_envio = adicionar_coluna_envio(pagina_clientes)

    ja_enviados = False

    # Verifica se há algum valor na coluna de envio
    for linha in pagina_clientes.iter_rows(min_row=2):
        if linha[coluna_envio - 1].value is not None:  # Se a célula não for None
            ja_enviados = True
            break  # Sai do loop se já encontrar um enviado

    return ja_enviados, None  # Retorna o estado de envio e nenhuma mensagem de erro
