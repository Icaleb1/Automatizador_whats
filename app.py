import logging
import os
from tkinter import messagebox
import customtkinter as ctk
from tkinter import *
import openpyxl, xlrd
import pathlib
from openpyxl import Workbook
from modulos.interface import carregar_arquivo, carregar_anexo
from modulos.automacaoWeb import inicializar_navegador, processar_clientes
from modulos.manipularLogs import configurar_logging, adicionar_email_handler
from verificarVersao import compararVersoes

#Setando a aparencia padrão do sistema
ctk.set_appearance_mode('System')
ctk.set_default_color_theme('blue')

cor_laranja = '#F25430'

arquivo_excel = None
#arquivo_excel_nome = ctk.StringVar('Nenhum arquivo selecionado.')


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.arquivo_excel_nome = ctk.StringVar(self, 'Nenhum arquivo selecionado')
        self.anexo_nome = ctk.StringVar(self, 'Nenhum anexo selecionado')
        self.mensagem = ctk.StringVar(self, '')
        self.arquivo_excel = None
        self.arquivo_anexo = None
        self.paginas = []
        self.lista_mensagens = []
        self.navegador = None

        configurar_logging()
        adicionar_email_handler()
        compararVersoes()

        self.layout_config()
        self.appearence()
        self.todo_sistema()

    def layout_config(self):
        self.title('Sistema de automação')
        self.geometry('800x600')
        self.maxsize(width=800, height=600)

    def appearence(self):
        self.lb_apm = ctk.CTkLabel(self, text='Tema', bg_color='transparent',
        text_color=['#000', '#fff']).place(x=50, y=500)
        self.opt_apm = (ctk.CTkOptionMenu(self,fg_color=cor_laranja, button_color=cor_laranja, button_hover_color=cor_laranja, 
        values=['Light', 'Dark', 'System'], command=self.change_apm).place
        (x=50, y=530))

    def change_apm(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)

    def todo_sistema(self):
      
        self.num_mensagens = IntVar()
        
        def selecionando_arquivo():
            self.arquivo_excel = carregar_arquivo()
            arquivo_excel = self.arquivo_excel
            if arquivo_excel:
                self.arquivo_excel_nome.set(pathlib.Path(arquivo_excel).name)
                try:
                    workbook = openpyxl.load_workbook(arquivo_excel)
                    paginas = workbook.sheetnames
                    paginas_combobox.configure(values=paginas)
                except Exception as e:
                    messagebox.showerror('Erro', f'Falha ao carregar o arquivo: {str(e)}')
    
        def selecionando_anexo():
                self.arquivo_anexo = carregar_anexo()
                arquivo_anexo = self.arquivo_anexo
                if arquivo_anexo:
                    self.anexo_nome.set(pathlib.Path(arquivo_anexo).name)

        def adicionar_mensagem():
            self.mensagem = self.entrada_mensagem.get("1.0", END).strip()  # Pega o texto da primeira linha até o final e remove espaços
            if self.mensagem:  # Verifica se a mensagem não está vazia
                self.lista_mensagens.append(self.mensagem)
                self.entrada_mensagem.delete("1.0", END)  # Limpa a caixa de texto após adicionar a mensagem
                messagebox.showinfo('Sucesso', 'Mensagem adicionada com sucesso!')
                atualizar_mensagem()
            else:
                messagebox.showwarning('Aviso', 'Mensagem não pode estar vazia!')

        def remover_mensagem():
            mensagem = self.mensagens_combobox.get()
            if mensagem in self.lista_mensagens:
                self.lista_mensagens.remove(mensagem)
                atualizar_mensagem()
            else:
                messagebox.showwarning('Aviso', 'Mensagem não encontrada!')

        def atualizar_mensagem():
            mensagens_truncadas = [msg[:30] + '...' 
            if len(msg) > 30
            else msg for msg in self.lista_mensagens]
            self.mensagens_combobox.configure(values=mensagens_truncadas)

        def selecionar_mensagem(event):
            indice = self.mensagens_combobox.current()
            if indice >= 0:
                mensagem_completa = self.lista_mensagens[indice]
                self.entrada_mensagem.delete('1.0', END) 
                self.entrada_mensagem.insert('1.0', mensagem_completa)

        def verificar_variaveis(mensagens, pagina_selecionada, workbook, file_path): 
        

            if not mensagens or not isinstance(mensagens, list):
                messagebox.showerror("Erro", "Mensagens estão vazias ou incorretas.")
                return False

            if not pagina_selecionada or pagina_selecionada == "Página não selecionada":
                messagebox.showerror("Erro", "Página da planilha de clientes está vazia ou não carregada.")
                return False

            if not workbook:
                messagebox.showerror("Erro", "Erro ao carregar o arquivo Excel.")
                return False
            
            if not file_path or not os.path.exists(file_path):
                messagebox.showerror("Erro", "Caminho do arquivo inválido ou inexistente.")
                return False

            return True

        def processar_mensagens():
            if self.lista_mensagens == None or self.arquivo_excel == None:
                    messagebox.showwarning('Sistema', 'Campos do formulário não preenchidos!')
            else:
                workbook = openpyxl.load_workbook(self.arquivo_excel)
                pagina_selecionada = paginas_combobox.get()

                if verificar_variaveis(self.lista_mensagens, pagina_selecionada, workbook, self.arquivo_excel):
                    try:
                        mensagens = self.lista_mensagens
                        workbook = openpyxl.load_workbook(self.arquivo_excel)
                        pagina = workbook[paginas_combobox.get()] 
                        navegador = inicializar_navegador()
                        processar_clientes(navegador, mensagens, pagina, workbook, self.arquivo_excel, self.arquivo_anexo)
                        messagebox.showinfo('Concluído', 'Mensagens enviadas.')
                    except Exception as e:
                        logging.error(f'Erro na função principal: {e}')
                    finally:
                        try:
                            navegador.quit()
                            limpar_formulario()
                        except Exception as e:
                            logging.error(f'Erro ao tentar fechar o navegador: {e}')
                        else:
                            open('erro.log', 'w').close()
                else:
                    pass

        def limpar_formulario():
            self.entrada_mensagem.delete("1.0", END)  # Limpa o campo de mensagem
            self.lista_mensagens = []  # Limpa a lista de mensagens
            self.mensagens_combobox.set('Selecione uma mensagem')  # Reseta o ComboBox de mensagens
            paginas_combobox.set('Página não selecionada')  # Reseta o ComboBox de páginas
            self.arquivo_excel = None  # Reseta o arquivo Excel
            self.arquivo_anexo = None  # Reseta o arquivo anexo
            self.arquivo_excel_nome.set('Nenhum arquivo selecionado')  # Reseta o nome do arquivo Excel
            self.anexo_nome.set('Nenhum anexo selecionado')  # Reseta o nome do anexo

            messagebox.showinfo('Sucesso', 'Formulário limpo com sucesso!')  # Mostra uma mensagem informando que o formulário foi limpo


        frame = (ctk.CTkFrame(self, width=800, height=50, corner_radius=0,
                bg_color='teal', fg_color=cor_laranja))
        frame.place(x=0, y=10)
        

        title = ctk.CTkLabel(frame, text='FlashLeads - Automação de processos', 
        font=('Century Gothic bold', 24),
        text_color='#fff')
        title.place(x=200, y=10)

        span = ctk.CTkLabel(self, text='Por Favor, preencha o formulário!', font=('Century Gothic bold', 16),
        text_color=['#000', '#fff']).place(x=50, y=70)

        frame_borda_arquivo = ctk.CTkFrame(self, width=280, height=40, fg_color='transparent',
        border_color=cor_laranja, border_width=2)
        frame_borda_arquivo.place(x=50, y=120)

        label_nome_arquivo = ctk.CTkLabel(frame_borda_arquivo, textvariable=self.arquivo_excel_nome, 
        font=('Century Gothic bold', 16),
        text_color=['#000', '#fff'], fg_color='transparent')
        label_nome_arquivo.place(x=5, y=5)

        botao_anexar_excel = ctk.CTkButton(self, text='Selecionar arquivo'.upper(),
        fg_color=cor_laranja, hover_color='#F7775B', command=selecionando_arquivo)
        botao_anexar_excel.place(x=50, y=170)

        #label_paginas = ctk.CTkLabel(self, text='Páginas do Excel:', 
        #font=('Century Gothic bold', 16),
        #text_color=['#000', '#fff'])
        #label_paginas.place(x=350, y=150)

        paginas_combobox = ctk.CTkComboBox(self, values=self.paginas, 
        font=('Century Gothic bold', 16),
        border_color=cor_laranja,
        button_color=cor_laranja, 
        width=305, height=40)
        paginas_combobox.set('Página não selecionada')
        paginas_combobox.place(x=400, y=120)

        # label_num_mensagens = ctk.CTkLabel(self, text='Número de mensagens',
        # font=('Century Gothic bold', 16),
        # text_color=['#000', '#fff'])
        # label_num_mensagens.place(x=50, y=250)

        # numero_mensagens = ctk.CTkEntry(self, width=100, 
        # textvariable=self.num_mensagens,
        # font=('Century Gothic bold', 16),
        # border_color=cor_laranja,
        # fg_color='transparent')
        # numero_mensagens.place(x=50, y=280)

        frame_borda_anexo = ctk.CTkFrame(self, width=280, height=40, fg_color='transparent',
        border_color=cor_laranja, border_width=2)
        frame_borda_anexo.place(x=50, y=230)

        label_nome_anexo = ctk.CTkLabel(frame_borda_anexo, textvariable=self.anexo_nome, 
        font=('Century Gothic bold', 16),
        text_color=['#000', '#fff'], fg_color='transparent')
        label_nome_anexo.place(x=5, y=5)

        botao_enviar_anexo = ctk.CTkButton(self, text='Selecionar anexo'.upper(),
        fg_color=cor_laranja, hover_color='#F7775B', command=selecionando_anexo)
        botao_enviar_anexo.place(x=50, y=280)

        label_mensagem = ctk.CTkLabel(self, text='Digite a mensagem:', 
        font=('Century Gothic bold', 16),
        text_color=['#000', '#fff'], fg_color='transparent')
        label_mensagem.place(x=50, y=345)

        self.entrada_mensagem = ctk.CTkTextbox(self, width=505 , height=150,
        font=('arial', 18), border_color=cor_laranja, border_width=2, fg_color='transparent')
        self.entrada_mensagem.place(x=200, y=350)

        self.mensagens_combobox = ctk.CTkComboBox(self, values=[], 
        font=('Century Gothic bold', 16),
        border_color=cor_laranja,
        button_color=cor_laranja, 
        width=305, height=40)
        self.mensagens_combobox.set('Selecione uma mensagem')
        self.mensagens_combobox.place(x=400, y=230)
        self.mensagens_combobox.bind('<<ComboboxSelected>>', selecionar_mensagem)
        
        botao_adicionar_msg_lista = ctk.CTkButton(self, text='Adicionar Mensagem'.upper(),
        fg_color=cor_laranja, hover_color='#F7775B', command=adicionar_mensagem)
        botao_adicionar_msg_lista.place(x=400, y=280)

        botao_remover_msg_lista = ctk.CTkButton(self, text='Remover Mensagem'.upper(),
        fg_color=cor_laranja, hover_color='#F7775B', command=remover_mensagem)
        botao_remover_msg_lista.place(x=560, y=280)

        botao_processar_mensagens = ctk.CTkButton(self, text='Processar Mensagens'.upper(),
        fg_color=cor_laranja, hover_color='#F7775B', command=processar_mensagens)
        botao_processar_mensagens.place(x=540, y=530)



if __name__=='__main__':
    app = App()
    app.mainloop()
